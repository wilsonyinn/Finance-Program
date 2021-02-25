import csv
import datetime
import openpyxl

wb = openpyxl.load_workbook('Finance Tracker.xlsx')
ws = wb['Sheet1']


with open('stmt.csv', newline='') as csvfile:
    reader = csv.reader(csvfile, delimiter=' ', quotechar='|')
    row_count = 1
    for row in reader:
        if row_count == 2:
            starting_balance = row[4][row[4].index('"') + 1:len(row[4]) - 1]
            starting_balance = float(starting_balance)
            start_date = row[4][0:10]
        if row_count == 3:
            totalDeposit = row[1][row[1].index('"') + 1: len(row[1]) - 1]
            totalDeposit = float(totalDeposit)
        if row_count == 4:
            totalWithdraw = row[1][row[1].index('"') + 1: len(row[1]) - 1]
            totalWithdraw = float(totalWithdraw)
        if row_count == 5:
            ending_date = row[4][0:10]
        if row_count >= 9:
            ws.append([row[0]])
            print(row)
        row_count += 1


ending_balance = starting_balance + totalDeposit + totalWithdraw

month_number = start_date[0:2]


datetime_object = datetime.datetime.strptime(month_number, "%m")
month_name = datetime_object.strftime("%B")
"""I came to the conclusion that the program is working and there are changes being made 
   but i do not know where to find the file. Two possibilities I can currently hypothesize:
                1) The changes are being made and I am actually writing on the spreadsheet, but I just cant find which one
                2) The changes just arent permanent/visible and i need to figure out how to make it permanent/visible
                """
print()
print(month_name + " Statement")
print()
print("Start Date: " + start_date)
print("End Date: " + ending_date)
print("Starting Balance: " + str(starting_balance))
print("Deposits: " + str(totalDeposit))
print("Withdrawls: " + str(totalWithdraw))
print("Ending Balance: " + str(ending_balance))


print(ws['C5'].value)
print(ws['A1'].value)
print(ws['F2'].value)
